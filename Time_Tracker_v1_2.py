import tkinter as tk
from tkinter import messagebox
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

class TimeTrackerApp(tk.Tk):
    """
    A simple time tracking application with a GUI.

    This app allows a user to enter a task name, start and stop a timer,
    and saves the time spent on each task to a daily Excel file with
    a new tab for each day. It also displays the total time spent on
    each task for the current day and includes daily summary calculations
    with a dynamic lunch deduction.
    """
    def __init__(self):
        super().__init__()

        print("--- Time Tracker Application Started ---")
        self.excel_filename = "TimeTrackerData.xlsx"
        print(f"Excel data will be saved to: {self.excel_filename}")
        print("Please ensure you have 'openpyxl' installed (pip install openpyxl).")

        # --- Window Configuration ---
        self.title("Advanced Time Tracker")
        self.geometry("500x450") # Slightly larger to accommodate new lunch options
        self.resizable(False, False)
        self.configure(bg="#f0f0f0")

        # --- State Variables ---
        self.timer_running = False
        self.start_time = None
        self.current_task = ""
        self.total_times = {} # Stores {task_name: total_seconds_for_task} for the current day
        self.elapsed_seconds = 0
        self.lunch_taken_var = tk.BooleanVar(self) # Variable for the lunch checkbox
        self.lunch_duration_minutes = 0 # Stores the calculated lunch duration in minutes

        # --- UI Element Creation ---
        self.create_widgets()

        # --- Initial Setup ---
        self.load_daily_data()
        self.update_totals_display()

    def create_widgets(self):
        """Creates and places all the GUI widgets."""
        main_frame = tk.Frame(self, bg="#f0f0f0", padx=20, pady=20)
        main_frame.pack(expand=True)

        # Task Input
        task_label = tk.Label(main_frame, text="Task Name:", font=("Helvetica", 12), bg="#f0f0f0")
        task_label.pack(pady=(0, 5))
        self.task_entry = tk.Entry(main_frame, width=45, font=("Helvetica", 12))
        self.task_entry.pack(pady=(0, 10))

        # Timer Display
        self.timer_display = tk.Label(main_frame, text="00:00:00", font=("Helvetica", 28, "bold"), bg="#f0f0f0")
        self.timer_display.pack(pady=(10, 20))

        # Control Buttons
        button_frame = tk.Frame(main_frame, bg="#f0f0f0")
        button_frame.pack(pady=10)
        self.start_button = tk.Button(button_frame, text="Start", command=self.start_timer, width=18, bg="#4CAF50", fg="white", font=("Helvetica", 12, "bold"), relief=tk.RAISED, bd=3)
        self.start_button.pack(side=tk.LEFT, padx=10)
        self.stop_button = tk.Button(button_frame, text="Stop", command=self.stop_timer, width=18, bg="#F44336", fg="white", font=("Helvetica", 12, "bold"), state=tk.DISABLED, relief=tk.RAISED, bd=3)
        self.stop_button.pack(side=tk.RIGHT, padx=10)

        # Lunch Options Frame
        lunch_frame = tk.Frame(main_frame, bg="#f0f0f0")
        lunch_frame.pack(pady=(15, 5))

        self.lunch_checkbox = tk.Checkbutton(lunch_frame, text="Took Lunch?", variable=self.lunch_taken_var,
                                             command=self.toggle_lunch_duration_entry,
                                             font=("Helvetica", 12), bg="#f0f0f0")
        self.lunch_checkbox.pack(side=tk.LEFT, padx=(0, 10))

        lunch_duration_label = tk.Label(lunch_frame, text="Duration (minutes):", font=("Helvetica", 12), bg="#f0f0f0")
        lunch_duration_label.pack(side=tk.LEFT)
        self.lunch_duration_entry = tk.Entry(lunch_frame, width=10, font=("Helvetica", 12), state=tk.DISABLED)
        self.lunch_duration_entry.pack(side=tk.LEFT)
        self.lunch_duration_entry.insert(0, "60") # Default to 60 minutes

        # Totals Display
        totals_label = tk.Label(main_frame, text="Today's Tracked Tasks:", font=("Helvetica", 12, "underline"), bg="#f0f0f0")
        totals_label.pack(pady=(15, 5))
        self.totals_text = tk.Text(main_frame, height=6, width=45, state=tk.DISABLED, bg="#ffffff", font=("Helvetica", 10), bd=1, relief="solid")
        self.totals_text.pack()

    def toggle_lunch_duration_entry(self):
        """Enables or disables the lunch duration entry based on checkbox state."""
        if self.lunch_taken_var.get():
            self.lunch_duration_entry.config(state=tk.NORMAL)
        else:
            self.lunch_duration_entry.config(state=tk.DISABLED)

    def get_sheet_name(self):
        """Generates the sheet name based on the current date."""
        return datetime.now().strftime("%Y-%m-%d")

    def load_daily_data(self):
        """
        Loads data from the current day's Excel sheet and updates totals.
        This ensures the in-memory `total_times` reflects the current day's data.
        """
        sheet_name = self.get_sheet_name()
        print(f"Attempting to load data from Excel sheet: '{sheet_name}' in '{self.excel_filename}'")
        self.total_times = {} # Reset totals for the current day

        if os.path.exists(self.excel_filename):
            try:
                workbook = load_workbook(self.excel_filename)
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    # Iterate through rows, skipping header and summary rows
                    for row_idx in range(2, sheet.max_row + 1): # Start from row 2 (after header)
                        task_name = sheet.cell(row=row_idx, column=1).value
                        duration_seconds = sheet.cell(row=row_idx, column=2).value
                        if task_name and isinstance(duration_seconds, (int, float)):
                            # Only load actual task entries, not summary rows
                            if task_name not in ["Total Seconds", "Total Minutes", "Total Hours", "Net Hours (after lunch)"]:
                                self.total_times[task_name] = self.total_times.get(task_name, 0) + int(duration_seconds)
                print(f"Successfully loaded existing data for '{sheet_name}'.")
            except Exception as e:
                messagebox.showerror("Error", f"Could not read Excel file: {e}")
                print(f"Error loading Excel file: {e}")
        else:
            print(f"Excel file '{self.excel_filename}' not found. A new one will be created.")

    def start_timer(self):
        """Starts the timer for the task entered by the user."""
        task_name = self.task_entry.get().strip()
        if not task_name:
            messagebox.showwarning("Input Error", "Please enter a task name.")
            return

        if self.timer_running:
            messagebox.showwarning("Timer Running", "A timer is already active. Please stop it first.")
            return

        print(f"Starting timer for task: '{task_name}'")
        self.current_task = task_name
        self.start_time = datetime.now()
        self.timer_running = True
        self.start_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.task_entry.config(state=tk.DISABLED)
        # Disable lunch options while timer is running
        self.lunch_checkbox.config(state=tk.DISABLED)
        self.lunch_duration_entry.config(state=tk.DISABLED)


        self.elapsed_seconds = 0
        self.update_timer_display()

    def stop_timer(self):
        """Stops the current timer, saves the data to Excel, and updates the display."""
        if not self.timer_running:
            messagebox.showwarning("No Timer", "No timer is currently running.")
            return

        end_time = datetime.now()
        elapsed_time = end_time - self.start_time
        elapsed_seconds = int(elapsed_time.total_seconds())

        print(f"Stopping timer for task: '{self.current_task}'. Elapsed time: {elapsed_seconds} seconds.")

        self.timer_running = False
        self.start_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)
        self.task_entry.config(state=tk.NORMAL)
        # Re-enable lunch options
        self.lunch_checkbox.config(state=tk.NORMAL)
        self.toggle_lunch_duration_entry() # Set state based on checkbox

        # Update total time for the task in memory
        self.total_times[self.current_task] = self.total_times.get(self.current_task, 0) + elapsed_seconds

        # Get lunch deduction amount
        lunch_deduction_seconds = 0
        if self.lunch_taken_var.get():
            try:
                lunch_minutes = float(self.lunch_duration_entry.get().strip())
                if lunch_minutes < 0:
                    messagebox.showwarning("Invalid Input", "Lunch duration cannot be negative. Setting to 0.")
                    lunch_minutes = 0
                lunch_deduction_seconds = int(lunch_minutes * 60)
                self.lunch_duration_minutes = lunch_minutes # Store for display if needed
            except ValueError:
                messagebox.showerror("Input Error", "Please enter a valid number for lunch duration (minutes). Setting to 0.")
                lunch_deduction_seconds = 0
                self.lunch_duration_minutes = 0
        else:
            self.lunch_duration_minutes = 0

        # Save the record and update totals in the Excel file
        self.save_to_excel(lunch_deduction_seconds)

        # Update the GUI totals display
        self.update_totals_display()

        self.current_task = ""
        self.timer_display.config(text="00:00:00")
        self.elapsed_seconds = 0

    def update_timer_display(self):
        """Updates the timer display every second."""
        if self.timer_running:
            self.elapsed_seconds = int((datetime.now() - self.start_time).total_seconds())
            hours, remainder = divmod(self.elapsed_seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            time_string = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            self.timer_display.config(text=time_string)
            self.after(1000, self.update_timer_display)

    def update_totals_display(self):
        """Updates the text widget with the total time for each task for the current day."""
        self.totals_text.config(state=tk.NORMAL)
        self.totals_text.delete("1.0", tk.END)
        if not self.total_times:
            self.totals_text.insert(tk.END, "No tasks tracked today.")
        else:
            for task, total_seconds in self.total_times.items():
                hours, remainder = divmod(total_seconds, 3600)
                minutes, seconds = divmod(remainder, 60)
                time_string = f"{hours:02d}h {minutes:02d}m {seconds:02d}s"
                self.totals_text.insert(tk.END, f"{task}: {time_string}\n")
        self.totals_text.config(state=tk.DISABLED)

    def save_to_excel(self, lunch_deduction_seconds):
        """
        Saves all tracked tasks for the current day to an Excel sheet,
        including calculated totals and dynamic lunch deduction.
        """
        sheet_name = self.get_sheet_name()
        print(f"Saving data to Excel sheet: '{sheet_name}' with lunch deduction of {lunch_deduction_seconds} seconds.")

        workbook = None
        try:
            if os.path.exists(self.excel_filename):
                workbook = load_workbook(self.excel_filename)
            else:
                workbook = Workbook()
                # Remove default 'Sheet' if it's empty and we're creating a new workbook
                if 'Sheet' in workbook.sheetnames and len(workbook['Sheet']._cells) == 0:
                    workbook.remove(workbook['Sheet'])

            if sheet_name not in workbook.sheetnames:
                sheet = workbook.create_sheet(sheet_name)
            else:
                sheet = workbook[sheet_name]
                # Clear existing content of the sheet to rewrite with updated totals
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.value = None

            # --- Write Headers ---
            headers = ["Task Name", "Seconds", "Minutes"]
            sheet.append(headers)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            for col_idx, header_text in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_idx, value=header_text)
                cell.font = header_font
                cell.fill = header_fill

            # --- Write Task Data ---
            current_row = 2
            daily_total_seconds = 0
            for task, total_seconds in self.total_times.items():
                total_minutes = total_seconds / 60
                sheet.cell(row=current_row, column=1, value=task)
                sheet.cell(row=current_row, column=2, value=total_seconds)
                # Apply number format for minutes
                minute_cell = sheet.cell(row=current_row, column=3, value=total_minutes)
                minute_cell.number_format = '0.00'
                daily_total_seconds += total_seconds

                # Apply row coloring based on task (simple alternating or specific)
                if current_row % 2 == 0:
                    row_fill = PatternFill(start_color="E0EBF5", end_color="E0EBF5", fill_type="solid") # Light blue
                else:
                    row_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid") # Light grey
                for col_idx in range(1, len(headers) + 1):
                    sheet.cell(row=current_row, column=col_idx).fill = row_fill
                current_row += 1

            # --- Write Summary Calculations ---
            daily_total_minutes = daily_total_seconds / 60
            daily_total_hours = daily_total_minutes / 60

            net_hours_after_lunch = daily_total_hours - (lunch_deduction_seconds / 3600) # Deduct dynamic lunch

            # Add a blank row for separation
            current_row += 1
            
            # Summary row for Total Seconds
            sheet.cell(row=current_row, column=1, value="Total Seconds")
            sheet.cell(row=current_row, column=2, value=daily_total_seconds)
            # Apply number format for total minutes
            total_minutes_cell_1 = sheet.cell(row=current_row, column=3, value=daily_total_minutes) # Display total minutes here too for clarity
            total_minutes_cell_1.number_format = '0.00'
            summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Light purple/blue
            for col_idx in range(1, len(headers) + 1):
                sheet.cell(row=current_row, column=col_idx).fill = summary_fill
            current_row += 1

            # Summary row for Total Minutes/Hours
            sheet.cell(row=current_row, column=1, value="Total Minutes")
            # Apply number format for total minutes
            total_minutes_cell_2 = sheet.cell(row=current_row, column=2, value=daily_total_minutes)
            total_minutes_cell_2.number_format = '0.00'
            # Apply number format for total hours
            total_hours_cell = sheet.cell(row=current_row, column=3, value=daily_total_hours) # Display total hours here
            total_hours_cell.number_format = '0.00'
            for col_idx in range(1, len(headers) + 1):
                sheet.cell(row=current_row, column=col_idx).fill = summary_fill
            current_row += 1

            # Summary row for Net Hours (after dynamic lunch)
            lunch_label = f"Net Hours (after {self.lunch_duration_minutes}min lunch)" if self.lunch_taken_var.get() else "Net Hours (no lunch deducted)"
            sheet.cell(row=current_row, column=1, value=lunch_label)
            sheet.cell(row=current_row, column=2, value=net_hours_after_lunch * 3600) # Convert back to seconds for consistency
            # Apply number format for net hours
            net_hours_cell = sheet.cell(row=current_row, column=3, value=net_hours_after_lunch)
            net_hours_cell.number_format = '0.00'
            net_hours_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Light orange/yellow
            net_hours_font = Font(bold=True)
            for col_idx in range(1, len(headers) + 1):
                cell = sheet.cell(row=current_row, column=col_idx)
                cell.fill = net_hours_fill
                cell.font = net_hours_font
            current_row += 1

            # --- Adjust Column Widths ---
            for col_idx in range(1, len(headers) + 1):
                sheet.column_dimensions[get_column_letter(col_idx)].width = 20

            workbook.save(self.excel_filename)
            print(f"Data successfully saved to '{self.excel_filename}' on sheet '{sheet_name}'.")

        except Exception as e:
            messagebox.showerror("Excel Save Error", f"Failed to save data to Excel: {e}")
            print(f"Error saving to Excel: {e}")
        finally:
            if workbook:
                try:
                    # Attempt to close the workbook if it was opened/created
                    # (openpyxl doesn't have an explicit close, but saving handles it)
                    pass
                except Exception as e:
                    print(f"Error during workbook cleanup: {e}")


if __name__ == "__main__":
    app = TimeTrackerApp()
    app.mainloop()
